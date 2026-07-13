"""
parsers/menu_mix.py
--------------------
Parses the Menu Mix report — item-level sales with categories.

Handles THREE format variants found in the client's 12 months of data:

  v1 (xlsx, Jul 2025 - Mar 2026):
      Header row 11, column "Super Category"
      Items grouped under category section headers
      Category name appears once per section, item rows below have
      a blank Super Category cell

  v2 (xlsx, Apr 2026+, e.g. "Menu_Mix_SCategory_By_all_all..."):
      Header row 9, column "SCategory" (renamed)
      Same section-header structure as v1, shifted up 2 rows
      Typo: "Disount Amount" instead of "Discount Amount"

  v3 (csv, Apr 2026 — "Enterprise_Menu_Mix_Report..."):
      Flat tabular format, no metadata headers
      Columns: Deployment, Category, Super Category, Item No., Item,
               Rate, Qty, Amount, Discount, Net Amount, Percentage of Sales
      Stops at "Grand Total" row in Deployment column

All three variants stop their item-level data at a "Grand Total" row —
after that is a category-summary section we don't need.

Output: a single standardised DataFrame regardless of input format:
    columns = [item_name, category, qty, amount, discount,
               net_amount, percentage_of_sales, item_name_lower]
"""

import pandas as pd
import openpyxl


def _detect_xlsx_version(ws) -> str:
    """
    Detects whether an xlsx Menu Mix file is v1 (old) or v2 (new format).

    v1: "Super Category" column header appears in row 12 (1-indexed)
    v2: "SCategory" column header appears in row 10 (1-indexed)
    """
    # Scan the first 15 rows for either header — more robust than
    # assuming an exact row number, since metadata row counts can drift
    for row_idx in range(1, 16):
        row_values = [
            str(c.value).strip() if c.value is not None else ""
            for c in ws[row_idx]
        ]
        if "Super Category" in row_values:
            return "v1"
        if "SCategory" in row_values:
            return "v2"

    raise ValueError(
        "Could not detect Menu Mix format version — "
        "neither 'Super Category' nor 'SCategory' header found in first 15 rows"
    )


def _parse_xlsx(filepath: str) -> pd.DataFrame:
    """
    Parses a Menu Mix xlsx file (v1 or v2 format).

    Both versions share the same row pattern:
        - One header row with column names
        - Items grouped under category section-header rows
          (category name in col 0, rest of row blank)
        - Item rows have BLANK category cell — category must be
          forward-filled from the section header above it
        - Stops at "Grand Total" row
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    version = _detect_xlsx_version(ws)

    all_rows = list(ws.iter_rows(values_only=True))

    # ── Find the header row ─────────────────────────────────────────────
    category_col_name = "Super Category" if version == "v1" else "SCategory"

    header_row_idx = None
    for i, row in enumerate(all_rows):
        if row and row[0] == category_col_name:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(f"Header row not found for {version} format")

    headers = [str(h).strip() if h else "" for h in all_rows[header_row_idx]]

    # ── Walk item rows, forward-filling category from section headers ──
    records = []
    current_category = None

    for row in all_rows[header_row_idx + 1:]:
        first_cell = row[0]

        # Stop at Grand Total — everything after is the summary section
        if first_cell == "Grand Total":
            break

        # Section header row: category name in col 0, rest blank
        # (col 1 "Number" is None for these rows)
        if first_cell is not None and row[1] is None:
            current_category = str(first_cell).strip()
            continue

        # Skip fully blank rows
        if first_cell is None and row[1] is None:
            continue

        # This is an item data row — build a dict using header names
        row_dict = dict(zip(headers, row))
        row_dict[category_col_name] = current_category
        records.append(row_dict)

    wb.close()

    df = pd.DataFrame(records)

    # ── Standardise column names across v1/v2 ───────────────────────────
    # v1 uses "Discount Amount", v2 has a typo "Disount Amount"
    rename_map = {
        category_col_name: "category",
        "Item Name":        "item_name",
        "Number Sold":      "qty",
        "Amount":           "amount",
        "Discount Amount":  "discount",
        "Disount Amount":   "discount",   # v2 typo
        "Net Sales":        "net_amount",
        "%  Of Sales":      "percentage_of_sales",
        "% Of Sales":       "percentage_of_sales",
    }
    df = df.rename(columns=rename_map)

    return df


def _parse_csv(filepath: str) -> pd.DataFrame:
    """
    Parses the Enterprise Menu Mix CSV format (v3).
    This is a flat tabular file — no metadata headers, no forward-fill
    needed. Stops at the "Grand Total" row in the Deployment column.
    """
    df = pd.read_csv(filepath)

    grand_total_mask = df["Deployment"].astype(str).str.strip() == "Grand Total"
    if grand_total_mask.any():
        first_total_idx = grand_total_mask.idxmax()
        df = df.loc[:first_total_idx - 1].copy()

    df = df.rename(columns={
        "Item":       "item_name",
        "Qty":        "qty",
        "Net Amount": "net_amount",
        "Percentage of Sales": "percentage_of_sales",
        # "category", "amount", "discount" already match after lowercasing
        "Category":   "category",
        "Amount":     "amount",
        "Discount":   "discount",
    })

    return df


def parse_menu_mix(filepath: str, file_format: str) -> pd.DataFrame:
    """
    Main entry point — parses a Menu Mix report regardless of format.

    Args:
        filepath:    local path to the downloaded file (e.g. /tmp/menu_mix.xlsx)
        file_format: "xlsx" or "csv" — determined from the S3 object's
                     file extension by the caller (lambda_function.py)

    Returns:
        Standardised DataFrame with columns:
        item_name, category, qty, amount, discount,
        net_amount, percentage_of_sales, item_name_lower
    """
    if file_format == "xlsx":
        df = _parse_xlsx(filepath)
    elif file_format == "csv":
        df = _parse_csv(filepath)
    else:
        raise ValueError(f"Unsupported file_format for menu_mix: {file_format}")

    # ── Common cleanup, identical for all formats ───────────────────────
    keep_cols = [
        "item_name", "category", "qty", "amount",
        "discount", "net_amount", "percentage_of_sales"
    ]
    # Some columns may be missing depending on format quirks — guard for it
    for col in keep_cols:
        if col not in df.columns:
            df[col] = None

    df = df[keep_cols].copy()

    for col in ["qty", "amount", "discount", "net_amount", "percentage_of_sales"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # ── Enforce float64 for all numeric columns ────────────────────────
    # pyarrow infers int64 when all values happen to be whole numbers
    # (e.g. Apr 2026 CSV has qty=148, amount=64440 — no decimals).
    # This causes HIVE_PARTITION_SCHEMA_MISMATCH in Athena when other
    # months write float64 for the same columns.
    # Explicit cast ensures every month writes the same Parquet type.
    for col in ["qty", "amount", "discount", "net_amount", "percentage_of_sales"]:
        df[col] = df[col].astype(float)

    df["item_name"] = df["item_name"].astype(str).str.strip()
    df["category"]  = df["category"].astype(str).str.strip()

    # ── Normalise category spelling inconsistency across formats ───────
    # The CSV (v3) source uses "Kababs and Roasted" while xlsx (v1/v2)
    # sources use "Kabab and Roasted" (no 's'). Without this fix, Athena
    # GROUP BY queries would silently split April's data into a separate
    # category bucket from every other month.
    CATEGORY_ALIASES = {
        "kababs and roasted": "Kabab and Roasted",
    }
    df["category"] = df["category"].apply(
        lambda c: CATEGORY_ALIASES.get(c.lower(), c)
    )

    df["item_name_lower"] = df["item_name"].str.lower()

    df = df.dropna(subset=["item_name"])
    df = df[df["item_name"].str.strip() != ""]
    df = df[df["item_name"].str.lower() != "nan"]

    return df.reset_index(drop=True)
